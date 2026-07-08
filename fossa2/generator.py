from openpyxl.worksheet.datavalidation import DataValidation
from .components.grupa.grupa_model import GrupaNaglowek
from .components.grupa_podmiotow.grupa_podmiotow_model import GrupaPodmioty
from .components.ankieta.ankieta_model import AnkietaNaglowek
from .components.ankieta_pytania.ankieta_pytania_model import AnkietaPytania
import re
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from .config import INFINITY_DATE

def get_data_by_grupanaglowek_id(grupanaglowek_id, poczatek_roku, koniec_roku):
    grupa_naglowek = GrupaNaglowek.objects.get(id=grupanaglowek_id)

    grupa_podmioty = GrupaPodmioty.objects.filter(
        id_grupa=grupanaglowek_id,
        data_od__lte=koniec_roku,
        data_do__gte=INFINITY_DATE,
        id_podmiotu__isnull=False
    ).select_related('id_podmiotu')
    ankiety = AnkietaNaglowek.objects.filter(
        id_grupa=grupanaglowek_id,
        data_od__lte=koniec_roku,
        data_do__gte=INFINITY_DATE
    )

    result = {
        'grupa_naglowek': {'id': grupa_naglowek.id, 'nazwa': grupa_naglowek.nazwa},
        'podmioty': [
            {'id': gp.id_podmiotu.id, 'kod': gp.id_podmiotu.kod, 'nazwa': gp.id_podmiotu.nazwa}
            for gp in grupa_podmioty
        ],
        'ankiety': [],
    }

    for ankieta in ankiety:
        ankieta_pytania = AnkietaPytania.objects.filter(id_ankieta_naglowek=ankieta.id).select_related(
            'id_podzapytania__id_pytania__id_podbloku__id_bloku__id_obszaru'
        )

        ankieta_data = {
            'id': ankieta.id,
            'nazwa': ankieta.nazwa,
            'elementy_ankiety': [],
        }

        for ap in ankieta_pytania:
            pz, pyt, pb, bl, ob = ap.id_podzapytania, ap.id_podzapytania.id_pytania, ap.id_podzapytania.id_pytania.id_podbloku, ap.id_podzapytania.id_pytania.id_podbloku.id_bloku, ap.id_podzapytania.id_pytania.id_podbloku.id_bloku.id_obszaru
            ankieta_data['elementy_ankiety'].append({
                'obszar': {'id': ob.id, 'kod': ob.kod, 'nazwa': ob.nazwa},
                'blok': {'id': bl.id, 'kod': bl.kod, 'tresc': bl.tresc},
                'podblok': {'id': pb.id, 'kod': pb.kod, 'tresc': pb.tresc},
                'pytanie': {'id': pyt.id, 'kod': pyt.kod, 'tresc': pyt.tresc},
                'podzapytanie': {'id': pz.id, 'kod': pz.kod, 'tresc': pz.tresc, 'obligatoryjne': pz.obligatoryjne, 'slownik': pz.slownik.opcje if pz.slownik else None}
            })
        result['ankiety'].append(ankieta_data)

    return result


def transform_dict(original_dict):
    podmioty = original_dict.get('podmioty', [])
    ankiety_z_hierarchia = []
    ankiety = original_dict.get('ankiety', [])

    for survey in ankiety:
        elementy = survey.get('elementy_ankiety', [])
        hierarchy = {}
        for item in elementy:
            ob = item['obszar']
            if ob['kod'] not in hierarchy:
                hierarchy[ob['kod']] = {'info': ob, 'bloki': {}}

            bl = item['blok']
            curr_obs = hierarchy[ob['kod']]['bloki']
            if bl['kod'] not in curr_obs:
                curr_obs[bl['kod']] = {'info': bl, 'podbloki': {}}

            pb = item['podblok']
            curr_bl = curr_obs[bl['kod']]['podbloki']
            if pb['kod'] not in curr_bl:
                curr_bl[pb['kod']] = {'info': pb, 'pytania': {}}

            pyt = item['pytanie']
            curr_pb = curr_bl[pb['kod']]['pytania']
            if pyt['kod'] not in curr_pb:
                curr_pb[pyt['kod']] = {'info': pyt, 'podzapytania': []}

            curr_pb[pyt['kod']]['podzapytania'].append(item['podzapytanie'])

    for obs in hierarchy.values():
        for bl in obs['bloki'].values():
            for pb in bl['podbloki'].values():
                for pyt in pb['pytania'].values():
                    pyt['podzapytania'].sort(key=lambda x: natural_sort_key(x['kod']))

    ankiety_z_hierarchia.append({
        'survey_info': {'id': survey['id'], 'nazwa': survey['nazwa']},
        'hierarchia': hierarchy
    })

    return {
        'ankiety_z_hierarchia': ankiety_z_hierarchia,
        'podmioty': podmioty,
        'hierarchia': hierarchy
    }



def natural_sort_key(s):
    if s is None:
        return []
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]


def generuj(hierarchia, podmiot, nazwa_ankiety):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odpowiedzi"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    fill_header_row = PatternFill(start_color='365f91', end_color='365f91', fill_type='solid')
    fill_bl = PatternFill(start_color='8db4e2', end_color='8db4e2', fill_type='solid')
    fill_pb = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
    fill_pyt = PatternFill(start_color='d9e3f0', end_color='d9e3f0', fill_type='solid')
    fill_input = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')
    font_white_bold = Font(bold=True, color="FFFFFF")
    font_black_bold = Font(bold=True, color="000000")


    headers = ["Nr", "Treść", "Odpowiedź", "Obligatoryjne", "Uzasadnienie"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = font_white_bold
        cell.fill = fill_header_row
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row_idx = 4


    for obs_kod in sorted(hierarchia.keys(), key=natural_sort_key):
        obs_data = hierarchia[obs_kod]
        nazwa_obs = obs_data['info']['kod']

        for bl_kod in sorted(obs_data['bloki'].keys(), key=natural_sort_key):
            bl_data = obs_data['bloki'][bl_kod]
            for c in range(1, 6):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = fill_bl
                cell.border = thin_border
                cell.font = font_black_bold
            ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {bl_kod}")
            ws.cell(row=row_idx, column=2, value=bl_data['info']['tresc'])
            row_idx += 1

            for pb_kod in sorted(bl_data['podbloki'].keys(), key=natural_sort_key):
                pb_data = bl_data['podbloki'][pb_kod]
                for c in range(1, 6):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.fill = fill_pb
                    cell.border = thin_border
                    cell.font = font_black_bold
                ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {pb_kod}")
                ws.cell(row=row_idx, column=2, value=pb_data['info']['tresc']).alignment = Alignment(indent=0)
                row_idx += 1

                for py_kod in sorted(pb_data['pytania'].keys(), key=natural_sort_key):
                    py_data = pb_data['pytania'][py_kod]
                    for c in range(1, 6):
                        cell = ws.cell(row=row_idx, column=c)
                        cell.fill = fill_pyt
                        cell.border = thin_border
                        cell.font = font_black_bold  # CZARNY
                    ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {py_kod}")
                    ws.cell(row=row_idx, column=2, value=py_data['info']['tresc']).alignment = Alignment(indent=0)
                    row_idx += 1

                    sorted_pz = sorted(py_data['podzapytania'], key=lambda x: natural_sort_key(x['kod']))
                    for pz in sorted_pz:
                        if pz['slownik']:
                            opcje = pz['slownik']
                            lista = ",".join(opcje)

                            dv = DataValidation(type="list", formula1=f'"{lista}"', allow_blank=True)
                            ws.add_data_validation(dv)


                            dv.add(ws.cell(row=row_idx, column=3))
                        for c in range(1, 6):
                            cell = ws.cell(row=row_idx, column=c)

                            cell.font = font_black_bold
                            cell.border = thin_border
                            if c in [3,4, 5]:
                                cell.fill = fill_input

                        ws.cell(row=row_idx, column=1, value=f"{nazwa_obs} {pz['kod']}")
                        ws.cell(row=row_idx, column=2, value=pz['tresc']).alignment = Alignment(indent=0)
                        ws.cell(row=row_idx, column=4, value="TAK" if pz['obligatoryjne'] else "NIE").alignment = Alignment(horizontal='center')
                        row_idx += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.freeze_panes = "A4"
    ws_info = wb.create_sheet(title="Informacje")

    ws_info.cell(row=1, column=1, value="Nazwa podmiotu:").font = font_black_bold
    ws_info.cell(row=1, column=2, value=podmiot['nazwa'])

    ws_info.cell(row=2, column=1, value="Kod podmiotu:").font = font_black_bold
    ws_info.cell(row=2, column=2, value=podmiot['kod'])

    ws_info.cell(row=3, column=1, value="Osoba odpowiedzialna za ankietę:").font = font_black_bold
    ws_info.cell(row=3, column=2, value="")
    ws_info.column_dimensions['A'].width = 50
    ws_info.column_dimensions['B'].width = 75

    output_dir = os.path.join(".", "fossa2", "generated_reports")
    if not os.path.exists(output_dir): os.makedirs(output_dir)

    safe_survey_name = re.sub(r'[^\w\s-]', '', nazwa_ankiety).strip().replace(' ', '_')
    filename = f"{datetime.datetime.now().year}_{podmiot['kod']}_{safe_survey_name}.xlsx"
    wb.save(os.path.join(output_dir, filename))

